#!/usr/bin/env python3
"""Create a 3-tab xlsx from extracted book chapter data."""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Style constants ──────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
ALT_ROW_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
EXPLICIT_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # amber
SUGGESTED_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")  # blue


def style_header(ws, cols):
    for col_idx, (label, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def style_row(ws, row_idx, num_cols, alt=False):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.alignment = CELL_ALIGN
        cell.border = THIN_BORDER
        if alt:
            cell.fill = ALT_ROW_FILL


def ch_sort_key(ch_str):
    """Sort 'Chapter N' strings numerically."""
    import re
    m = re.search(r'(\d+)', ch_str)
    return int(m.group(1)) if m else 999


# ── DATA ─────────────────────────────────────────────────────────────────────

citations = [
    # ── Chapter 1 ──
    ("Chapter 1", "The current landscape", "[Insert articles about usage and Gen AI pilots not taking off]", "Explicit", "Placeholder for articles about Gen AI usage and pilots not gaining traction"),
    ("Chapter 1", "Why is change happening: the people", "[examples of AI training programs]", "Explicit", "Placeholder for examples of AI training programs at law firms"),
    ("Chapter 1", "Why is change happening: the people", "[10000 prompts example]", "Explicit", "Placeholder for 10,000 prompts usage-driving initiative example"),
    ("Chapter 1", "Why is change happening: the people", "firms making AI usage part of the appraisal [example] and even the bonus [example]", "Explicit", "Two placeholders for examples of firms tying AI usage to appraisals and bonuses"),
    ("Chapter 1", "Why change is happening: competition", "alternative business structures (ABSs) have been in existence since [ ]", "Explicit", "Blank date for when ABSs were introduced in England and Wales"),
    ("Chapter 1", "Why change is happening: competition", "\"the unauthorized practice of law may die with a whimper not a bang\" [Damien - need to ask for permission]", "Explicit", "Attributed quote from 'Damien' — needs permission and proper citation"),
    ("Chapter 1", "Time, transparency, and the metrics problem", "law firm comparison sites [Jamie Tso example] where clients can choose the firm based on the best price", "Explicit", "Placeholder for Jamie Tso example about law firm comparison sites"),
    ("Chapter 1", "The AI landscape in law", "Natural Language Processing (NLP) in the early 2000s enabled 'understanding' text... Machine Learning (ML) in the mid-2010s brought patterns recognition", "Suggested", "Historical claims about NLP and ML timelines in legal tech could benefit from citations"),
    ("Chapter 1", "A snapshot: SKILLS Gen AI Use Cases Survey", "The SKILLS Questionnaire provides a snapshot of how legal professionals currently use technology", "Suggested", "The SKILLS survey should be formally cited with year, methodology, and sample size"),
    ("Chapter 1", "Why change is happening: competition", "Many of these legal technology AI companies have multi-million dollar valuations", "Suggested", "Claim about legal tech AI company valuations should be supported with data"),
    ("Chapter 1", "Why change is happening: competition", "one observer calling these vendors 'the biggest Trojan Horse in legal'", "Suggested", "Attributed quote to unnamed observer — needs attribution"),
    ("Chapter 1", "Why is change happening: the economic reality", "Alternative fee arrangements have become more common, but are still far from the norm.", "Suggested", "Claim about AFA prevalence could benefit from survey data"),
    ("Chapter 1", "Why is change happening: the economic reality", "some clients are becoming more sophisticated in how they buy legal services, requiring secondments or innovation projects", "Suggested", "Specific client practices would benefit from an example or citation"),
    # ── Chapter 2 ──
    ("Chapter 2", "What is vibe coding?", "[Source] — when Claude announced a Claude Legal plug-in. It affected the shares of legal technology companies.", "Explicit", "Placeholder for share price impact of Claude Legal plug-in announcement"),
    ("Chapter 2", "Model Context Protocol (MCP)", "APIs ([define])", "Explicit", "Placeholder to define APIs"),
    ("Chapter 2", "What is vibe coding?", "This topic is reviewed in more detail in Chapter [ ]", "Explicit", "Blank chapter reference — needs correct chapter number for vibe coding discussion"),
    ("Chapter 2", "Probabilistic vs. Deterministic", "Catherine Bamford of Bam!Legal [request permission] states that lawyers will use a combination of generative AI and more traditional deterministic automation", "Explicit", "Attributed quote from Catherine Bamford — needs permission and formal citation"),
    ("Chapter 2", "A brief history of artificial intelligence", "Researchers have been attempting to build machines that mimic human reasoning for more than seventy years.", "Suggested", "Could cite founding AI works (e.g., Turing 1950, Dartmouth Conference 1956)"),
    ("Chapter 2", "A brief history of artificial intelligence", "interest in AI rose and fell in cycles often referred to as 'AI summers' and 'AI winters'", "Suggested", "AI winters/summers concept could benefit from a history-of-AI source"),
    ("Chapter 2", "Introduction", "It has become a common refrain that lawyers will not be replaced by AI, but will be replaced by another lawyer using AI.", "Suggested", "Widely attributed quote — should cite original source"),
    ("Chapter 2", "What is vibe coding?", "when Claude announced a Claude Legal plug-in. It affected the shares of a number of legal technology companies.", "Suggested", "Specific market impact claim needs a financial data source"),
    # ── Chapter 3 ──
    ("Chapter 3", "Pricing for value", "a small number of institutional clients have begun experimenting with credit-based fee structures [citation needed]", "Explicit", "Citation needed for credit-based fee structure examples in the UK"),
    ("Chapter 3", "Pricing for value", "Case Study: First Pass Due Diligence from Mayer Brown [Citation Needed] [Permission needed]", "Explicit", "Citation and permission needed for the Mayer Brown case study"),
    ("Chapter 3", "AI-enhanced pricing", "As Richard Susskind said, it is hard to tell a group of millionaires that their business model is wrong [citation needed]", "Explicit", "Citation needed for Richard Susskind quote"),
    ("Chapter 3", "AI-enhanced pricing", "[What tools are available for this?] [citation needed]", "Explicit", "Placeholder for tools available for AI-enhanced pricing"),
    ("Chapter 3", "Charging for knowledge, not technology", "a client developed a bespoke tool for withholding tax review and asked all of its panel firms to use it [citation needed]", "Explicit", "Citation needed for the withholding tax review tool example"),
    ("Chapter 3", "Innovation as a route to new markets", "[Citations needed: Ropes & Gray DownMarket Diligence?; Paul Weiss building workflows with Harvey; Allen & Overy's Antitrust workflow]", "Explicit", "Citations needed for three specific firm AI innovation examples"),
    ("Chapter 3", "The billable hour, revisited", "Ropes & Gray now credits first and second year associates with up to 400 hours of AI-related work [citation needed]", "Explicit", "Citation needed for the Ropes & Gray 400 AI hours policy"),
    ("Chapter 3", "Why there cannot be a standalone AI strategy", "some categories of work are susceptible to Jevons Paradox [citation needed]", "Explicit", "Citation needed for Jevons Paradox applied to legal AI"),
    ("Chapter 3", "Pricing for value", "AFAs and EFAs tend to be more prevalent in transactional practices in Europe and Asia", "Suggested", "Claim about geographic AFA/EFA prevalence should be supported with market data"),
    ("Chapter 3", "Pricing for value", "In the States, especially in the BigLaw market, outcome based pricing is not yet the norm.", "Suggested", "US BigLaw pricing claim could benefit from survey or report citation"),
    ("Chapter 3", "Charging for knowledge, not technology", "In eDiscovery, clients historically paid third-party vendors directly. When law firms brought eDiscovery in-house, clients continued to pay", "Suggested", "Historical eDiscovery pricing claim could benefit from an industry source"),
    # ── Chapter 4 ──
    ("Chapter 4", "Introduction", "[cite to Adam's book? - briefly quote his discussion of T-shaped people?]", "Explicit", "Potential citation to 'Adam's book' about T-shaped people"),
    ("Chapter 4", "Knowledge and innovation: one department", "A senior candidate refused the split [citation needed? Permission needed]", "Explicit", "Citation/permission needed for anecdote about a candidate refusing a split role"),
    ("Chapter 4", "A note on the shifting technology landscape", "legal quant [define in footnote with citation]", "Explicit", "The term 'legal quant' needs a definition in a footnote with citation"),
    ("Chapter 4", "Data scientists and analysts", "data scientists built a model to detect when time was being billed to the wrong matter [citation needed]", "Explicit", "Citation needed for the billing error detection model anecdote"),
    ("Chapter 4", "A note on the shifting technology landscape", "the release of Claude Opus 4.5 in November of 2025 and the democratization of SaaS development is changing the calculus", "Suggested", "Specific product release claim should be verified and cited"),
    ("Chapter 4", "A note on the shifting technology landscape", "The first wave of legal AI consisted of point solutions like Kira and Luminance... The second wave brought end-to-end workbench tools", "Suggested", "The three-wave legal AI framework would benefit from citation if drawn from existing taxonomy"),
    ("Chapter 4", "Where does AI live?", "There is a strong consensus, borne out across firms of all sizes and geographies, that AI capability should not sit primarily within IT.", "Suggested", "Claim about industry consensus could benefit from survey or report citation"),
    # ── Chapter 5 ──
    ("Chapter 5", "Data: the new oil?", "Investment in legal technology is higher than ever. [citation needed]", "Explicit", "Citation needed for legal technology investment levels claim"),
    ("Chapter 5", "MCP: Changing the legal data landscape", "[Major legal AI providers like Harvey partnered with companies like Lexis and FromCounsel...]", "Explicit", "Bracketed factual claim about partnerships and product launches requiring citation"),
    ("Chapter 5", "Citations section #1", "SKILLS community survey data on time savings (e.g., 11% savings in research time) — verify specific figures and publication details.", "Explicit", "Numbered citation item: SKILLS survey time savings data"),
    ("Chapter 5", "Citations section #2", "Artificial Lawyer article on the 'architecture of intelligence' — locate specific article.", "Explicit", "Numbered citation item: Artificial Lawyer article reference"),
    ("Chapter 5", "Citations section #3", "The Matrix (1999) — film reference for the end-scene analogy.", "Explicit", "Numbered citation item: film reference"),
    ("Chapter 5", "Citations section #4", "Examples of meta-dashboards in other professional services (investment banking, consulting, accounting).", "Explicit", "Numbered citation item: professional services dashboard examples"),
    ("Chapter 5", "Citations section #5", "SKILLS survey data on financial data and metrics — verify what was reported.", "Explicit", "Numbered citation item: SKILLS financial data"),
    ("Chapter 5", "Entering The Matrix", "In other professional services industries it is routine. Investment banks and management consultancies routinely analyse engagements in this way.", "Suggested", "Comparative claim about investment banks needs a specific example or source"),
    ("Chapter 5", "MCP: Changing the legal data landscape", "These companies did not simply build search engines... they became the single points of truth for the profession.", "Suggested", "Claim about legal research companies' historical role could benefit from a market analysis"),
    # ── Chapter 6 ──
    ("Chapter 6", "Engagement and Adoption: The Incentive Problem", "Ropes & Gray credits first-year associates with up to four hundred hours of AI-related work [CITATION NEEDED: Ropes & Gray AI credit programme details]", "Explicit", "Citation needed for Ropes & Gray AI credit programme"),
    ("Chapter 6", "Citations section #1", "Ropes & Gray AI credit programme — 400 hours for first-year associates. Verify current programme details.", "Explicit", "Numbered citation item: Ropes & Gray programme"),
    ("Chapter 6", "Citations section #2", "Mayer Brown associate adoption anecdote — associates would not use AI until hitting billable targets. [PERMISSION NEEDED]", "Explicit", "Numbered citation item: Mayer Brown anecdote, needs permission"),
    ("Chapter 6", "Citations section #3", "Harvey and Legora law school programmes — verify current programmes and specific schools.", "Explicit", "Numbered citation item: vendor law school training programmes"),
    ("Chapter 6", "Citations section #4", "UK training contract/apprenticeship model — verify accuracy.", "Explicit", "Numbered citation item: UK legal training model"),
    ("Chapter 6", "Citations section #5", "Investment banking analyst/apprenticeship model — verify accuracy.", "Explicit", "Numbered citation item: banking analogy verification"),
    ("Chapter 6", "Citations section #6", "Clio-Vlex acquisition — verify deal details and timing.", "Explicit", "Numbered citation item: Clio-Vlex M&A details"),
    ("Chapter 6", "Citations section #7", "'Implementing AI in Your Business' — locate full citation.", "Explicit", "Numbered citation item: external source for adoption framework"),
    ("Chapter 6", "Citations section #8", "SKILLS survey data on engagement and adoption metrics.", "Explicit", "Numbered citation item: SKILLS survey on adoption"),
    ("Chapter 6", "Citations section #9", "Current highest reported hourly billing rate ($4,000) — verify source.", "Explicit", "Numbered citation item: $4,000/hour rate"),
    ("Chapter 6", "Citations section #10", "Client anecdote on bringing AI-generated drafts to law firms. [PERMISSION NEEDED]", "Explicit", "Numbered citation item: client AI draft anecdote"),
    ("Chapter 6", "Law Schools", "Harvey and Legora are taking their rivalry to law schools, offering tool-specific training.", "Suggested", "Specific vendor law school claims need examples of programmes or schools"),
    ("Chapter 6", "Measuring ROI", "The highest reported hourly rate in the market at the time of writing is approximately four thousand dollars.", "Suggested", "The $4,000/hour rate claim needs a verifiable source"),
    ("Chapter 6", "Engagement and Adoption", "One firm reported that after a thorough rollout, associates would not use it until they had hit their billable hour targets.", "Suggested", "Anonymised firm anecdote — clarify if this is Mayer Brown or keep anonymous"),
    ("Chapter 6", "Teaching Judgment, Not Just Technology", "The Clio-Vlex acquisition is an expression of exactly this logic.", "Suggested", "References a specific acquisition — needs deal date/value citation"),
    # ── Chapter 7 ──
    ("Chapter 7", "The State of the Market", "Legal Tech Hub and similar directories catalogue and review products across the market [citation needed].", "Explicit", "Citation needed for Legal Tech Hub reference"),
    ("Chapter 7", "The State of the Market", "The SKILLS community publishes data on who is using what and how they rate it [Citation needed: SKILLS community publications]", "Explicit", "Citation needed for SKILLS community data"),
    ("Chapter 7", "Prioritization", "[Citation needed: Oz's client prioritization framework, shared in zoom discussions]", "Explicit", "Citation needed for prioritisation framework"),
    ("Chapter 7", "Conflicts of Interest in Tool Selection", "Some firms are beginning to address this [citation needed?]", "Explicit", "Citation needed for firms with AI vendor conflict-of-interest policies"),
    ("Chapter 7", "Benchmarking", "community resources like the SKILLS survey [citation needed] provide ongoing data on tool performance.", "Explicit", "Citation needed for SKILLS survey benchmarking data"),
    ("Chapter 7", "The Update Problem", "[Citation needed; permission needed?: Harvey third-party web search provider change]", "Explicit", "Citation needed for Harvey changing a third-party provider"),
    ("Chapter 7", "Security and Data Governance in Procurement", "[Citation needed: Adam's firm SOC 2 and ISO certification]", "Explicit", "Citation needed for a firm's SOC 2 and ISO certification"),
    ("Chapter 7", "The State of the Market", "The legal technology market is moving at a pace that has no precedent. New products launch weekly.", "Suggested", "Market pace claim could be supported by a market report"),
    ("Chapter 7", "One Platform or Many Tools?", "Orbital Witness, with its connections to land registries and title databases, outperforms a general platform for property due diligence.", "Suggested", "Product capability claim could cite product documentation or case study"),
    ("Chapter 7", "Security and Data Governance", "firms currently invest significant effort in ensuring their AI tools respect ethical walls, mapping iManage workspaces", "Suggested", "Industry practice claim could reference a case study or vendor documentation"),
    ("Chapter 7", "One Platform or Many Tools?", "Disputes teams cannot route all eDiscovery work through a general-purpose AI tool if the firm also uses Relativity or Everlaw", "Suggested", "Specific product limitations claim could cite product documentation"),
    # ── Chapter 8 ──
    ("Chapter 8", "Introduction / Risk Framework", "the legal profession has a well-documented bias toward the status quo and risk aversion [citation needed]", "Explicit", "Citation needed for legal profession's documented risk aversion"),
    ("Chapter 8", "What Keeps People Paralyzed", "Economists call it Jevons Paradox: when a resource becomes more efficient to use, people use more of it [citation needed]", "Explicit", "Citation needed for Jevons Paradox reference"),
    ("Chapter 8", "What Keeps People Paralyzed", "United States v. Heppner and Warner v. Gilbarco, Inc. on AI and privilege are only the most prominent examples [citation needed]", "Explicit", "Citation needed for Heppner and Warner decisions"),
    ("Chapter 8", "Competence and responsibility", "ABA Model Rule 1.1 requires lawyers to provide competent representation, including keeping abreast of relevant technology [citation needed]", "Explicit", "Citation needed for ABA Model Rule 1.1 and Comment 8"),
    ("Chapter 8", "Competence and responsibility", "Emerging regulations in several jurisdictions require organizations to assess and address bias in AI systems [citation needed]", "Explicit", "Citation needed for AI bias regulations (e.g., EU AI Act, NYC Local Law 144)"),
    ("Chapter 8", "Competence and responsibility", "some firms, particularly in the UK, have begun to consider energy consumption as part of their responsible AI framework [citation needed]", "Explicit", "Citation needed for UK firms' energy consumption AI frameworks"),
    ("Chapter 8", "Competence and responsibility", "Some firms have restricted certain high-consumption features, such as AI-generated video [citation needed]", "Explicit", "Citation needed for firms restricting high-consumption AI features"),
    ("Chapter 8", "Client confidentiality", "There is a documented instance of Google indexing shareable ChatGPT links [citation needed]", "Explicit", "Citation needed for Google indexing ChatGPT links incident"),
    ("Chapter 8", "Client confidentiality", "The democratization of AI agents has outpaced most lawyers' understanding of what they are consenting to [citation needed]", "Explicit", "Citation needed for AI agent tool access risks"),
    ("Chapter 8", "Client confidentiality", "governance tools which provide an anonymization layer, scrubbing identifying information [citation needed]", "Explicit", "Citation needed for anonymization/governance tools for AI workflows"),
    ("Chapter 8", "Client confidentiality", "on-premise or private-cloud deployments with local language models [citation needed]", "Explicit", "Citation needed for on-premise LLM deployment examples in law firms"),
    ("Chapter 8", "Attorney-Client Privilege", "In United States v. Heppner, [citation needed] No. 25-cr-00503 (S.D.N.Y., February 2026)", "Explicit", "Citation needed for U.S. v. Heppner decision"),
    ("Chapter 8", "Attorney-Client Privilege", "in Warner v. Gilbarco, Inc. [citation needed] (E.D. Mich., February 2026)", "Explicit", "Citation needed for Warner v. Gilbarco decision"),
    ("Chapter 8", "Attorney-Client Privilege", "disclosing information to ChatGPT did not constitute a waiver [citation needed] [Cleary Gottlieb memo, Feb 2026]", "Explicit", "Citation needed for waiver standard; Cleary Gottlieb memo referenced but not fully cited"),
    ("Chapter 8", "Talking to Clients About AI", "The 2026 SKILLS community survey data shows a significant proportion of clients have some form of restriction [citation needed]", "Explicit", "Citation needed for 2026 SKILLS survey on client AI restrictions"),
    ("Chapter 8", "What Keeps People Paralyzed", "a lawyer who uses AI may return a thirty-six-page markup where a manual review might have produced eight pages", "Suggested", "Specific numerical claim (36 vs 8 pages) — clarify if hypothetical or cite source"),
    # ── Chapter 9 ──
    ("Chapter 9", "Introduction", "The term vibe coding originates with AI researcher Andrej Karpathy [CITATION NEEDED: Andrej Karpathy, original vibe coding post]", "Explicit", "Citation needed for Karpathy's original vibe coding post/tweet"),
    ("Chapter 9", "Recognising and Rewarding Innovation", "Ropes and Gray's programme granting first-year associates four hundred hours of AI credit [CITATION NEEDED]", "Explicit", "Citation needed for Ropes & Gray AI credit programme"),
    ("Chapter 9", "Citations section #1", "Andrej Karpathy — original 'vibe coding' tweet or post. Verify source and exact phrasing.", "Explicit", "Numbered citation item: Karpathy vibe coding origin"),
    ("Chapter 9", "Citations section #2", "Base44 acquisition by Wix — verify $80 million figure and timeline (six months, one-person company).", "Explicit", "Numbered citation item: Base44/Wix acquisition"),
    ("Chapter 9", "Citations section #3", "Claude Code and Claude Cowork — Anthropic product details and current capabilities.", "Explicit", "Numbered citation item: Anthropic product details"),
    ("Chapter 9", "Citations section #4", "Manus — verify product details and status.", "Explicit", "Numbered citation item: Manus product verification"),
    ("Chapter 9", "Citations section #5", "Ropes & Gray AI credit programme — 400 hours. Verify details and whether public.", "Explicit", "Numbered citation item: Ropes & Gray programme"),
    ("Chapter 9", "Citations section #6", "Jamie's products — determine whether to reference by name or anonymise.", "Explicit", "Numbered citation item: Jamie attribution decision"),
    ("Chapter 9", "Citations section #7", "OpenClaw security incidents — verify agent security risk details.", "Explicit", "Numbered citation item: OpenClaw security risk"),
    ("Chapter 9", "Citations section #8", "Thomson Reuters Institute or similar reports on shadow AI prevalence in law firms.", "Explicit", "Numbered citation item: shadow AI prevalence data"),
    ("Chapter 9", "Citations section #9", "ABA survey data on solo practitioner AI adoption.", "Explicit", "Numbered citation item: ABA solo practitioner survey"),
    ("Chapter 9", "Citations section #10", "eDiscovery revenue data for large firms — for the cannibalization discussion.", "Explicit", "Numbered citation item: eDiscovery revenue data"),
    ("Chapter 9", "The Democratisation of Building", "Base44, acquired by Wix for $80 million after just six months of operation by a single developer", "Suggested", "Specific acquisition figure and timeline are factual claims requiring verification"),
    ("Chapter 9", "Shadow AI: The Invisible Adoption Curve", "This is happening at every firm. It is happening right now.", "Suggested", "Universal claim about shadow AI prevalence could benefit from survey data"),
    # ── Chapter 11 ──
    ("Chapter 11", "People and Networks", "Justin Levy has published a useful list of people to follow in this space *[citation needed]*", "Explicit", "Citation needed for Justin Levy's list of legal tech people to follow"),
    ("Chapter 11", "Communities", "SKILLS is a community oriented primarily toward law firms... *[citation needed]*", "Explicit", "Citation needed for SKILLS community description/website"),
    ("Chapter 11", "Communities", "CLOC and LegalOps serve a parallel function for in-house legal teams [*citation needed*]", "Explicit", "Citation needed for CLOC and LegalOps details"),
    ("Chapter 11", "Communities", "ILTA occupies a broader position in the ecosystem... *[citation needed]*", "Explicit", "Citation needed for ILTA description/website"),
    ("Chapter 11", "Communities", "informal communities, including LegalQuants, that operate outside the established conference circuit *[citation needed]*", "Explicit", "Citation needed for LegalQuants and regional K&I meetup groups"),
    ("Chapter 11", "The Role of Consultants", "The Legaltech Hub... operating as 'the gardener of legal tech' *[citation needed]*", "Explicit", "Citation needed for Legaltech Hub description and quote attribution"),
    ("Chapter 11", "Professional Bodies and Associations", "Many have established AI task forces or legal technology committees [*citation needed*]", "Explicit", "Citation needed for bar association AI task forces"),
    ("Chapter 11", "Professional Bodies and Associations", "the Law Society has conducted surveys on the use of legal technology *[citation needed]*", "Explicit", "Citation needed for Law Society technology surveys"),
    ("Chapter 11", "Professional Bodies and Associations", "firms like Kraft Kennedy and Harbor Global offer managed IT and technology consulting [*citation needed*]", "Explicit", "Citation needed for Kraft Kennedy and Harbor Global services"),
]

cross_references = [
    # ── Chapter 1 ──
    ("Chapter 1", "Chapter 2", "For a more in-depth review of the history and workings of AI, see Chapter 2.", "Explicit"),
    # ── Chapter 2 ──
    ("Chapter 2", "Chapter 1", "As was defined in Chapter 1, 'agents' are autonomous software systems powered by generative AI", "Explicit"),
    ("Chapter 2", "Unknown", "This topic is reviewed in more detail in Chapter [ ] (vibe coding)", "Explicit"),
    ("Chapter 2", "Chapter 3", "As we will explore in the following chapter, the more important challenge becomes strategic", "Implicit"),
    ("Chapter 2", "Chapter 3", "For Later Chapters: ROI measurement (ch 3)", "Flagged"),
    ("Chapter 2", "Chapter 4", "For Later Chapters: Where AI lives organizationally, changing role of innovation teams (ch 4?)", "Flagged"),
    ("Chapter 2", "Chapter 6", "For Later Chapters: AI training as a top priority (ch 6)", "Flagged"),
    # ── Chapter 3 ──
    ("Chapter 3", "Chapter 5", "This is also, critically, another argument for the data strategy discussed in [Chapter 5].", "Explicit"),
    ("Chapter 3", "Chapter 6", "Much like with training [see Chapter [6]]", "Explicit"),
    ("Chapter 3", "Later chapters", "The chapters that follow will examine this in detail: the skills, the training, the governance", "Implicit"),
    # ── Chapter 4 ──
    ("Chapter 4", "Chapters 1-2", "Chapters 1 and 2 set out what generative AI is, why it matters to the legal profession", "Explicit"),
    ("Chapter 4", "Chapter 3", "Chapter 3 explored the (potentially existential) strategic consequences of this technology.", "Explicit"),
    ("Chapter 4", "Chapter 3", "The strategic questions we explored in Chapter 3, about pricing, hiring, leverage", "Explicit"),
    ("Chapter 4", "Chapter 3", "as discussed in Chapter 3, new lines of business", "Explicit"),
    ("Chapter 4", "Chapter 7", "there is a chapter in this book dedicated to tool selection [see Chapter 7]", "Explicit"),
    ("Chapter 4", "Chapters 7, 9", "We return to this in detail in Chapter 7 and Chapter 9.", "Explicit"),
    ("Chapter 4", "Chapter 9", "[see chapter 9] — the emerging legal quant", "Explicit"),
    ("Chapter 4", "Chapter 8", "This is discussed further in the context of governance in Chapter [ 8 ]", "Explicit"),
    ("Chapter 4", "Chapter 3", "As discussed in the previous chapter, clients are asking deeper questions.", "Implicit"),
    ("Chapter 4", "Chapter 8", "For later chapters: AI governance frameworks, risk policies (Chapter 8)", "Flagged"),
    ("Chapter 4", "Chapters 3, 8", "For later chapters: Monitoring and measuring AI usage (Chapters 3 and 8)", "Flagged"),
    ("Chapter 4", "Chapter 6", "For later chapters: Change management, communications strategy (Chapter 6)", "Flagged"),
    ("Chapter 4", "Chapter 7", "For later chapters: Tool selection, UX/UI (Chapter 7)", "Flagged"),
    ("Chapter 4", "Chapter 9", "For later chapters: Shadow AI and vibe coding (ch 9)", "Flagged"),
    ("Chapter 4", "Chapter 8", "For later chapters: Ethical walls, security models (Chapter 8)", "Flagged"),
    ("Chapter 4", "Chapters 6, 9", "For later chapters: Evolution of law school training (Chapter 6 and potentially Chapter 9)", "Flagged"),
    ("Chapter 4", "Chapter 9", "For later chapters: Firms as tech companies (ch 9)", "Flagged"),
    # ── Chapter 5 ──
    ("Chapter 5", "Chapter 3", "Flagged: ROI measurement problem connects to pricing/business model discussion in Chapter 3.", "Flagged"),
    ("Chapter 5", "Chapter 7", "Flagged: Artificial Lawyer article may be relevant to tool evaluation in Chapter 7.", "Flagged"),
    ("Chapter 5", "Chapter 8", "Flagged: Cascading error risk with agentic AI has implications for Chapter 8 risk governance.", "Flagged"),
    # ── Chapter 6 ──
    ("Chapter 6", "Chapter 3", "This connects directly to the business model questions discussed in Chapter 3.", "Explicit"),
    ("Chapter 6", "Chapter 3", "Flagged: ROI discussion connects to pricing conversation in Chapter 3.", "Flagged"),
    ("Chapter 6", "Chapter 4", "Flagged: T-shaped lawyers, vibe coding lawyers, legal engineers for Chapter 4.", "Flagged"),
    ("Chapter 6", "Chapter 7", "The next chapter turns to selecting the right tools within an increasingly crowded landscape.", "Implicit"),
    ("Chapter 6", "Chapter 7", "Flagged: FirmOS aspiration of Harvey/Legora for tool selection discussion in Chapter 7.", "Flagged"),
    ("Chapter 6", "Chapter 9", "Flagged: Teaching people to vibe code for Chapter 9.", "Flagged"),
    ("Chapter 6", "Chapter 10", "Flagged: Client AI conversation reframes the outside counsel relationship for Chapter 10.", "Flagged"),
    # ── Chapter 7 ──
    ("Chapter 7", "Earlier chapters", "If the previous chapters have established why generative AI matters...", "Implicit"),
    ("Chapter 7", "Chapter 11", "Conferences, peer networks, and industry groups all contribute. [see chapter 11].", "Explicit"),
    ("Chapter 7", "Chapter 4", "As discussed in [Chapter 4], an AI committee is not optional.", "Explicit"),
    ("Chapter 7", "Chapter 3", "[see also, discussion about this in chapter 3]", "Explicit"),
    ("Chapter 7", "Chapter 4", "As discussed briefly in [Chapter 4], the traditional debate was build versus buy.", "Explicit"),
    ("Chapter 7", "Chapter 9", "This is explored in detail in [Chapter 9], but it warrants mention here.", "Explicit"),
    ("Chapter 7", "Chapter 6", "As discussed in [Chapter 6], training is a continuous process.", "Explicit"),
    ("Chapter 7", "Chapter 8", "Security and data governance, explored more fully in [Chapter 8].", "Explicit"),
    ("Chapter 7", "Chapter 9", "This will be explored further in [Chapter 9].", "Explicit"),
    ("Chapter 7", "Chapter 11", "Rely on external resources, directories, peer networks [See chapter 11].", "Explicit"),
    ("Chapter 7", "Chapter 4", "it connects backward, to the team structures and skills discussed in [Chapter 4]", "Explicit"),
    ("Chapter 7", "Chapter 5", "the data foundations discussed in [Chapter 5]", "Explicit"),
    ("Chapter 7", "Chapter 6", "the training and adoption strategies discussed in [Chapter 6]", "Explicit"),
    # ── Chapter 8 ──
    ("Chapter 8", "Chapter 7", "As discussed in [Chapter 7], the tool selection landscape is evolving at unprecedented speed", "Explicit"),
    ("Chapter 8", "Chapter 3", "This dynamic and its pricing implications are discussed further in [Chapter 3]", "Explicit"),
    ("Chapter 8", "Chapter 4", "The roles of innovation teams in supporting client conversations are discussed in [Chapter 4]", "Explicit"),
    ("Chapter 8", "Chapter 3", "The pricing implications of this shift are explored in depth in [Chapter 3]", "Explicit"),
    ("Chapter 8", "Chapter 7", "This framework, discussed in detail in [Chapter 7] as preferred/permitted/prohibited", "Explicit"),
    ("Chapter 8", "Chapter 4", "as discussed in the context of the AI committee in [Chapter 4]", "Explicit"),
    ("Chapter 8", "Chapter 9", "The next chapter turns to vibe coding and shadow AI", "Implicit"),
    # ── Chapter 9 ──
    ("Chapter 9", "Chapter 8", "as discussed in the previous chapter, the proliferation of agent-based tools creates risk", "Implicit"),
    ("Chapter 9", "Chapter 7", "The rise of vibe coding forces a reconsideration of the build-versus-buy decision discussed in Chapter 7", "Explicit"),
    ("Chapter 9", "Chapter 7", "as discussed in Chapter 7, the answer is often to rent", "Explicit"),
    ("Chapter 9", "Chapter 10", "The next chapter turns to the partnerships between in-house legal teams, outside counsel, and vendors", "Implicit"),
    ("Chapter 9", "Chapter 4", "Flagged: Linklaters model of solicitors embedded in innovation teams — for Chapter 4.", "Flagged"),
    ("Chapter 9", "Chapter 4", "Flagged: Power users vs broader population; how firms identify/cultivate them — for Chapter 4.", "Flagged"),
    ("Chapter 9", "Chapter 7", "Flagged: Lawyer-built tools and the preferred/permitted/prohibited framework — for Chapter 7.", "Flagged"),
    ("Chapter 9", "Chapter 8", "Flagged: Security risks of agent-based tools like OpenClaw — for Chapter 8.", "Flagged"),
    ("Chapter 9", "Chapter 10", "Flagged: Vendor roadmap influence via internal prototypes — for Chapter 10.", "Flagged"),
    ("Chapter 9", "Chapter 10", "Flagged: eDiscovery revenue tension with vendors (Relativity, Everlaw) — for Chapter 10.", "Flagged"),
    # ── Chapter 11 ──
    ("Chapter 11", "Chapter 9", "The relevant skills and implications are discussed in [Chapter 9].", "Explicit"),
    ("Chapter 11", "Chapter 8", "[Chapter 8] discussed the risk framework for this kind of experimentation", "Explicit"),
]

# ── Index Terms ──────────────────────────────────────────────────────────────
# Merged across all chapters. Format: (term, chapters_str, context, category)
index_terms_raw = [
    # Concepts
    ("Generative AI", "1, 2, 3, 4, 5, 6, 7, 8, 9, 11", "This book is being written because with Generative AI something *has* changed", "Concept"),
    ("Agentic AI", "1, 2, 3, 4, 5, 7", "agentic AI and beyond promises autonomous, multi-step goal execution", "Concept"),
    ("Vibe coding", "1, 2, 4, 5, 6, 9, 11", "the practice of creating software through conversational prompts to AI rather than through traditional programming", "Concept"),
    ("Shadow AI", "1, 4, 6, 8, 9", "the use of AI tools outside of official firm channels", "Concept"),
    ("Hallucination", "2, 5, 8", "when the model generates information that appears plausible but is factually incorrect or entirely fabricated", "Concept"),
    ("Retrieval-augmented generation (RAG)", "2", "many systems use techniques such as retrieval-augmented generation (RAG)", "Concept"),
    ("Transformer", "2", "What transformers allowed these tools to do is to consider much more context all at once", "Concept"),
    ("Token", "2", "Large language models process language not as words but as tokens", "Concept"),
    ("Context window", "2", "the amount of text that the model can consider at one time", "Concept"),
    ("Attention mechanism", "2", "This mechanism is known as 'attention'", "Concept"),
    ("Model Context Protocol (MCP)", "2, 5", "a standardised way for AI models to connect to tools, applications, and data sources", "Concept"),
    ("Probabilistic vs deterministic reasoning", "2", "This process is stochastic, meaning it involves probabilities rather than deterministic rules", "Concept"),
    ("Neural network", "2", "The idea of mapping technology like a human brain is known as a 'neural network'", "Concept"),
    ("Prompt engineering", "2, 4", "the advent of generative AI shifted this question to whether lawyers should be prompt engineers", "Concept"),
    ("Value-based pricing", "1, 3", "seeks to find a compromise between the traditional hours model and the benefits to the client", "Concept"),
    ("Law as a Service (LaaS)", "1", "'law as a service' (LaaS)... where law is more akin to streaming services", "Concept"),
    ("Legal streaming", "1", "Law firms 'broadcast' their services into platform providers", "Concept"),
    ("Jevons Paradox", "3, 6, 8", "when a resource becomes more efficient to use, people use more of it, not less", "Concept"),
    ("Opportunity cost", "3", "the relevant metric is opportunity cost, not hours saved", "Concept"),
    ("Leverage model", "3", "It destabilizes long-standing assumptions about pricing, leverage, value", "Concept"),
    ("Data hygiene", "3, 4", "The firms that have invested in data hygiene will be able to price with confidence", "Concept"),
    ("Change management", "1, 3, 4, 6, 7", "Confidence drops when questions turn to process, data, technology, and change management", "Concept"),
    ("Digital transformation", "1, 3, 8, 9", "Technology is delivering on things previously overpromised in other hype cycles", "Concept"),
    ("Knowledge management", "4, 5, 11", "knowledge management in law firms has been organised around individual items", "Concept"),
    ("Innovation team", "1, 4, 9", "Technology within firms is usually managed by centralized innovation teams", "Concept"),
    ("AI committee", "4, 7, 8", "An AI committee is not optional. It is the mechanism through which competing priorities are managed", "Concept"),
    ("T-shaped people / T-shaped lawyer", "4, 6", "an elegant evolution of the T-shaped persona", "Concept"),
    ("Legal quant", "4", "a capable legal engineer or legal quant with the right foundation model and the right data", "Concept"),
    ("Legal engineer", "4, 5, 6", "professionals who understand both the technology and the legal context", "Concept"),
    ("Process mapping", "4, 7", "Process mapping is the foundation of workflow automation and agentic AI", "Concept"),
    ("Service design", "4", "understanding who the users are, what their needs are, and how to design technology-supported processes", "Concept"),
    ("Chief Innovation Officer", "4, 9", "The role of a chief innovation officer, at best, is that of a chief marginal improvement officer", "Concept"),
    ("Data as work product", "5", "data, understood properly, is the work product. Not an input to the work product.", "Concept"),
    ("Matter-level data", "5", "A legal team with clean, structured, matter-level data can build AI systems that understand how its lawyers practise", "Concept"),
    ("Cascading error risk", "5", "If the wrong information enters at the beginning, the consequences cascade through the system", "Concept"),
    ("Data literacy", "5", "Data literacy is no longer solely the domain of IT professionals", "Concept"),
    ("Anonymisation / synthetic datasets", "5", "make appropriate use of anonymisation, aggregation, and synthetic datasets", "Concept"),
    ("Adoption (technology)", "6, 7", "There is no innovation without adoption, and it takes a village", "Concept"),
    ("Competency-based progression", "6", "Instead of first year, second year, firms will increasingly adopt level 1, level 2, level 3", "Concept"),
    ("AI mentor", "6", "One of the most consistent themes in conversations with junior lawyers is the anxiety of asking questions", "Concept"),
    ("Learning-by-simulation", "6", "Firms can create simulation environments that replace learning-by-doing", "Concept"),
    ("Return on investment (ROI) in legal AI", "6", "The question of return on investment in AI tools is asked constantly and answered poorly", "Concept"),
    ("FirmOS / operating system of the law firm", "6, 7", "Harvey, Legora, and similar platforms aspire to become the operating system of the law firm", "Concept"),
    ("Local champions (adoption)", "6", "Successful adoption requires local champions who understand the culture of their practice group", "Concept"),
    ("Gamification", "6", "experimentation with different approaches: gamification for some groups, competition for others", "Concept"),
    ("Tool selection", "7", "how do you actually choose the right tools?", "Concept"),
    ("Technology overload", "7", "a particular kind of overload that is distinct from what law firms have experienced before", "Concept"),
    ("Platform vs point solution", "6, 7", "the difference between a platform play and a point solution", "Concept"),
    ("Prioritisation framework", "7", "The most effective frameworks assess potential applications against several dimensions", "Concept"),
    ("Conflicts of interest (tool selection)", "7", "Investment relationships between law firm partners and legal tech companies create a problem", "Concept"),
    ("Pilot design", "7", "Pilot design matters as much as the evaluation itself", "Concept"),
    ("Benchmarking (AI tools)", "7", "Benchmarking AI tools presents a challenge that did not exist with previous generations", "Concept"),
    ("Multi-model architecture", "7", "the shift to multi-model architectures has made model-level benchmarking less relevant", "Concept"),
    ("Preferred / permitted / prohibited framework", "7, 8, 9", "classify tools into three categories: preferred, permitted, and prohibited", "Concept"),
    ("Build, buy, or rent", "7, 9", "the traditional debate was build versus buy. Generative AI introduced a third option: rent.", "Concept"),
    ("Sandbox environment", "7, 9", "Testing features in a sandbox environment before they reach production is essential", "Concept"),
    ("Net promoter score (NPS)", "1, 7", "Do clients start to rank their firms with net promoter scores?", "Concept"),
    ("Bias in AI outputs", "8", "AI systems trained on historical data will reflect the biases present in that data", "Concept"),
    ("Data loss prevention (DLP)", "8", "Others deploy data loss prevention tools purpose-built for AI workflows", "Concept"),
    ("Anonymization layer", "8", "governance tools which provide an anonymization layer, scrubbing identifying information", "Concept"),
    ("AI use policy", "8", "the firm must establish a clear, comprehensive AI use policy", "Concept"),
    ("On-premise deployment", "8", "on-premise or private-cloud deployments with local language models", "Concept"),
    ("Risk aversion (legal profession)", "8", "the legal profession has a well-documented bias toward the status quo", "Concept"),
    ("Innovation incentives", "9", "creating innovation bonus pools, giving billable credit for time spent building tools", "Concept"),
    ("Cannibalization", "9", "what happens when internally built tools compete with the firm's own revenue-generating services?", "Concept"),
    ("Horizon scanning", "11", "Closely related to personal testing is the practice of seeing what others are building", "Concept"),
    ("Data governance", "5, 7, 8", "the quality, structure, and governance of legal data become professional responsibilities", "Concept"),
    ("Procurement (legal technology)", "7", "the procurement process determines whether the firm gets value from its investment", "Concept"),
    ("Data room", "7", "Data rooms were once considered risky... met with significant resistance", "Concept"),
    ("Law+ (philosophy)", "6", "the broader Law+ philosophy that the legal profession requires multidisciplinary awareness", "Concept"),

    # Legal Terms
    ("Billable hour", "1, 3, 4, 6, 8, 9", "Law firms bill by the hour", "Legal Term"),
    ("Alternative fee arrangements (AFAs)", "1, 3", "Alternative fee arrangements (AFAs) and effective fee arrangements (EFAs) are on the rise", "Legal Term"),
    ("Effective fee arrangements (EFAs)", "1, 3", "Alternative fee arrangements (AFAs) and effective fee arrangements (EFAs) are on the rise", "Legal Term"),
    ("Alternative business structure (ABS)", "1", "alternative business structures (ABSs) have been in existence since [ ]", "Legal Term"),
    ("Unauthorized practice of law", "1", "the unauthorized practice of law may die with a whimper not a bang", "Legal Term"),
    ("Due diligence", "1, 2, 3, 6, 7", "Let us take due diligence as a good example of the difference agents make", "Legal Term"),
    ("Document review", "1, 2, 3", "from research platforms to document review systems", "Legal Term"),
    ("eDiscovery", "3, 6, 7, 9", "In eDiscovery, clients historically paid third-party vendors directly", "Legal Term"),
    ("Playbook", "1, 3", "an AI platform designed to review a contract against a playbook", "Legal Term"),
    ("Profits per partner (PPP)", "3", "A firm that measures success solely by this year's PPP", "Legal Term"),
    ("Realization rate", "3", "what the effective realization rate was for similar matters", "Legal Term"),
    ("Write-down / write-off", "3", "firms absorbed the risk through write-downs and write-offs", "Legal Term"),
    ("Panel firm", "1, 3", "in-house teams can do the work that previously panel firms would have completed", "Legal Term"),
    ("Professional indemnity insurance", "1, 3", "paying for the security of the law firm's insurance should things go awry", "Legal Term"),
    ("Matter (legal)", "5", "law firms work in matters, not documents", "Legal Term"),
    ("Client confidentiality", "5, 8", "Law firms have considerable obligations relating to client confidentiality", "Legal Term"),
    ("Ethical walls", "5, 7", "There may be ethical walls in place in sensitive matters", "Legal Term"),
    ("Training contract (UK)", "6", "mandatory training contracts for newly qualified solicitors", "Legal Term"),
    ("Solicitors Qualifying Examination (SQE)", "6", "the introduction of the Solicitors Qualifying Examination (SQE)", "Legal Term"),
    ("ABA Model Rule 1.1", "8", "ABA Model Rule 1.1 requires lawyers to provide competent representation", "Legal Term"),
    ("Duty of competence", "2, 8", "keeping abreast of changes in the law including the benefits and risks associated with relevant technology", "Legal Term"),
    ("Attorney-client privilege", "8, 9", "risk that communications between lawyer and client lose their protected status when routed through AI", "Legal Term"),
    ("Privilege waiver", "8", "privilege addresses the risk of waiver, and the consequences of waiver can be catastrophic", "Legal Term"),
    ("Work-product doctrine", "8", "neither attorney-client privilege nor the work-product doctrine applied", "Legal Term"),
    ("Outside counsel guidelines (OCGs)", "8", "Clients are increasingly setting expectations about AI use through outside counsel guidelines", "Legal Term"),
    ("Continuing legal education (CLE)", "8", "CLE should evolve to include AI competence as a core component", "Legal Term"),
    ("Professional negligence", "8", "The practical risk is professional negligence", "Legal Term"),
    ("Cross-border data transfers", "8", "strict rules on cross-border data transfers", "Legal Term"),
    ("Data processing agreements", "8", "tools that operate under appropriate data processing agreements", "Legal Term"),
    ("General Counsel", "8, 11", "The General Counsel should have a meaningful role in development and oversight of the policy", "Legal Term"),
    ("In-house legal department", "1, 3, 4", "In-house legal departments have become increasingly sophisticated", "Legal Term"),
    ("Legal operations", "1, 3", "In-house legal teams are leveraging opportunities to implement legal operations teams", "Legal Term"),
    ("Subscription model", "1, 3", "the growing interest in subscription models, bespoke builds, and modular offerings", "Legal Term"),
    ("United States v. Heppner", "8", "Judge Rakoff held that documents generated using a publicly available AI tool were not privileged", "Legal Term"),
    ("Warner v. Gilbarco, Inc.", "8", "Magistrate Judge Patti reached the opposite conclusion on AI and privilege", "Legal Term"),

    # Products
    ("Harvey", "1, 2, 3, 4, 5, 6, 7, 9", "Major corporations deploying the same AI platforms as their outside counsel (such as Harvey)", "Product"),
    ("Legora", "1, 4, 6, 7, 9", "end-to-end workbench tools like Harvey and Legora", "Product"),
    ("Claude", "2, 4", "Claude — Anthropic's generative AI tool", "Product"),
    ("Claude Code", "4, 7, 9", "Claude Code, Anthropic's coding agent, can build applications through conversation", "Product"),
    ("Claude Cowork", "9", "Claude Cowork takes this further, removing even the need to understand code", "Product"),
    ("Claude Opus 4.5", "4", "the release of Claude Opus 4.5 in November of 2025", "Product"),
    ("ChatGPT", "4, 8, 9", "A lawyer uses a personal ChatGPT account to draft a memo", "Product"),
    ("Kira", "4", "The first wave of legal AI consisted of point solutions like Kira and Luminance", "Product"),
    ("Luminance", "4", "The first wave of legal AI consisted of point solutions like Kira and Luminance", "Product"),
    ("CoCounsel", "5", "the research companies built their own tools like CoCounsel and Protege", "Product"),
    ("Protege", "5", "the research companies built their own tools like CoCounsel and Protege", "Product"),
    ("Relativity", "6, 7, 9", "the firm also uses Relativity or Everlaw for that purpose", "Product"),
    ("Everlaw", "7, 9", "the firm also uses Relativity or Everlaw", "Product"),
    ("Orbital Witness", "7", "a specialised tool with its connections to land registries and title databases", "Product"),
    ("iManage", "7", "mapping iManage workspaces and matter numbers", "Product"),
    ("Clio", "6", "The Clio-Vlex acquisition is an expression of exactly this logic", "Product"),
    ("Vlex", "6", "The Clio-Vlex acquisition", "Product"),
    ("SharePoint", "5", "utilising tools of the rest of the business, like SharePoint", "Product"),
    ("Codex", "7", "using tools like Claude Code, Codex, or Lovable", "Product"),
    ("Lovable", "7", "using tools like Claude Code, Codex, or Lovable", "Product"),
    ("Base44", "9", "Base44, acquired by Wix for $80 million after just six months", "Product"),
    ("Manus", "9", "Manus and similar tools offer yet another path", "Product"),
    ("DeepJudge", "9", "the firm might implement within Legora, Harvey, DeepJudge", "Product"),
    ("OpenClaw", "9", "agent-based tools like OpenClaw granting broad system access", "Product"),
    ("LinkedIn", "11", "LinkedIn has become the de facto platform where legal tech professionals share", "Product"),
    ("Microsoft ecosystem", "11", "tools within the Microsoft ecosystem may already offer capabilities", "Product"),
    ("Siri", "2", "we anthropomorphise our tools by giving them names. Just like 'Siri'", "Product"),
    ("Cortana", "2", "we anthropomorphise our tools by giving them names... 'Cortana'", "Product"),

    # Organizations
    ("SKILLS (Strategic Knowledge & Innovation Legal Leaders' Summit)", "1, 5, 6, 7, 8, 11", "SKILLS community survey provides a snapshot of how legal professionals use technology", "Organization"),
    ("ILTA", "11", "ILTA serves law firms, in-house teams, and vendors", "Organization"),
    ("CLOC", "11", "CLOC and LegalOps serve a parallel function for in-house legal teams", "Organization"),
    ("LegalOps", "11", "CLOC and LegalOps serve a parallel function for in-house legal teams", "Organization"),
    ("LegalQuants", "9, 11", "newer groups like LegalQuants that operate outside the established conference circuit", "Organization"),
    ("Legaltech Hub", "7, 11", "The Legaltech Hub... operating as 'the gardener of legal tech'", "Organization"),
    ("Anthropic", "2, 4, 9", "Claude — Anthropic's generative AI tool", "Organization"),
    ("Ropes & Gray", "3, 6, 9", "Ropes & Gray now credits first-year associates with up to 400 hours of AI-related work", "Organization"),
    ("Allen & Overy", "3", "Allen & Overy's Antitrust workflow", "Organization"),
    ("Paul Weiss", "3", "Paul Weiss building workflows with Harvey", "Organization"),
    ("Mayer Brown", "3, 6", "Case Study: First Pass Due Diligence from Mayer Brown", "Organization"),
    ("Bam!Legal", "2", "Catherine Bamford of Bam!Legal", "Organization"),
    ("Linklaters", "9", "The Linklaters model of practicing solicitors embedded in innovation teams", "Organization"),
    ("Wix", "9", "Base44, acquired by Wix for $80 million", "Organization"),
    ("Thomson Reuters Institute", "9", "Thomson Reuters Institute or similar reports on shadow AI prevalence", "Organization"),
    ("Cleary Gottlieb", "8", "Cleary Gottlieb, Courts Grapple with Privilege Implications of AI, February 2026", "Organization"),
    ("Law Society (England and Wales)", "11", "the Law Society has conducted surveys on the use of legal technology", "Organization"),
    ("Kraft Kennedy", "11", "firms like Kraft Kennedy offer managed IT and technology consulting", "Organization"),
    ("Harbor Global", "11", "firms like Kraft Kennedy and Harbor Global", "Organization"),
    ("Artificial Lawyer", "5", "Artificial Lawyer article on the 'architecture of intelligence'", "Organization"),
    ("LexisNexis", "5", "Harvey partnered with companies like Lexis and FromCounsel", "Organization"),
    ("FromCounsel", "5", "Harvey partnered with companies like Lexis and FromCounsel", "Organization"),

    # Persons
    ("Richard Susskind", "3", "As Richard Susskind said, it is hard to tell a group of millionaires that their business model is wrong", "Person"),
    ("Catherine Bamford", "2", "Catherine Bamford of Bam!Legal states that lawyers will use a combination of approaches", "Person"),
    ("Andrej Karpathy", "9", "The term vibe coding originates with AI researcher Andrej Karpathy", "Person"),
    ("Justin Levy", "11", "Justin Levy has published a useful list of people to follow in this space", "Person"),

    # Acronyms
    ("AFA — Alternative Fee Arrangement", "1, 3", "Alternative fee arrangements (AFAs)", "Acronym"),
    ("EFA — Effective Fee Arrangement", "1, 3", "Effective fee arrangements (EFAs)", "Acronym"),
    ("ABS — Alternative Business Structure", "1", "alternative business structures (ABSs)", "Acronym"),
    ("LLM — Large Language Model", "1, 2, 8", "Large Language Models (LLMs)", "Acronym"),
    ("NLP — Natural Language Processing", "1, 2", "Natural Language Processing (NLP)", "Acronym"),
    ("ML — Machine Learning", "1", "Machine Learning (ML)", "Acronym"),
    ("MCP — Model Context Protocol", "2, 5", "Model Context Protocol (MCP)", "Acronym"),
    ("RAG — Retrieval-Augmented Generation", "2", "retrieval-augmented generation (RAG)", "Acronym"),
    ("KPI — Key Performance Indicator", "1, 4", "key performance indicator (KPI)", "Acronym"),
    ("PPP — Profits Per Partner", "3", "PPP (Profits Per Partner)", "Acronym"),
    ("SaaS — Software as a Service", "4", "the democratization of SaaS development", "Acronym"),
    ("RFP — Request for Proposal", "4", "Responding to RFPs with a checklist of AI tools", "Acronym"),
    ("ROI — Return on Investment", "6", "Measuring ROI: Short, Medium, and Long Term", "Acronym"),
    ("SQE — Solicitors Qualifying Examination", "6", "the introduction of the Solicitors Qualifying Examination (SQE)", "Acronym"),
    ("SOC 2", "7", "Some firms have pursued SOC 2 and ISO certification", "Acronym"),
    ("GDPR — General Data Protection Regulation", "8", "Regulations such as the GDPR and the CCPA", "Acronym"),
    ("CCPA — California Consumer Privacy Act", "8", "Regulations such as the GDPR and the CCPA", "Acronym"),
    ("ABA — American Bar Association", "8, 9, 11", "ABA Model Rule 1.1", "Acronym"),
    ("API — Application Programming Interface", "2", "APIs, the standard connectors that allowed earlier generations of software to integrate", "Acronym"),
    ("DLP — Data Loss Prevention", "8", "data loss prevention tools purpose-built for AI workflows", "Acronym"),
    ("ERP — Enterprise Resource Planning", "6", "the appeal is the same as an ERP", "Acronym"),
    ("OCG — Outside Counsel Guidelines", "8", "outside counsel guidelines", "Acronym"),
    ("NPS — Net Promoter Score", "1, 7", "net promoter scores", "Acronym"),
    ("LaaS — Law as a Service", "1", "law as a service (LaaS)", "Acronym"),
]

# ── Build the workbook ───────────────────────────────────────────────────────
wb = Workbook()

# ── Tab 1: Citations Needed ──
ws1 = wb.active
ws1.title = "Citations Needed"
ws1.sheet_properties.tabColor = "DC2626"
cols1 = [("Chapter", 14), ("Section", 32), ("Excerpt", 70), ("Type", 14), ("Notes", 50)]
style_header(ws1, cols1)

# Sort: by chapter (numerically), then Explicit before Suggested
citations.sort(key=lambda x: (ch_sort_key(x[0]), 0 if x[3] == "Explicit" else 1))

for i, (ch, section, excerpt, ctype, notes) in enumerate(citations, 2):
    ws1.cell(row=i, column=1, value=ch)
    ws1.cell(row=i, column=2, value=section)
    ws1.cell(row=i, column=3, value=excerpt)
    ws1.cell(row=i, column=4, value=ctype)
    ws1.cell(row=i, column=5, value=notes)
    style_row(ws1, i, 5, alt=(i % 2 == 0))
    # Color the type cell
    type_cell = ws1.cell(row=i, column=4)
    if ctype == "Explicit":
        type_cell.fill = EXPLICIT_FILL
    else:
        type_cell.fill = SUGGESTED_FILL

# ── Tab 2: Cross References ──
ws2 = wb.create_sheet("Cross References")
ws2.sheet_properties.tabColor = "2563EB"
cols2 = [("Source Chapter", 16), ("Target Chapter", 18), ("Excerpt", 75), ("Type", 14)]
style_header(ws2, cols2)

cross_references.sort(key=lambda x: (ch_sort_key(x[0]), ch_sort_key(x[1])))

for i, (src, tgt, excerpt, rtype) in enumerate(cross_references, 2):
    ws2.cell(row=i, column=1, value=src)
    ws2.cell(row=i, column=2, value=tgt)
    ws2.cell(row=i, column=3, value=excerpt)
    ws2.cell(row=i, column=4, value=rtype)
    style_row(ws2, i, 4, alt=(i % 2 == 0))

# ── Tab 3: Index Terms ──
ws3 = wb.create_sheet("Index Terms")
ws3.sheet_properties.tabColor = "16A34A"
cols3 = [("Term", 38), ("Chapter(s)", 22), ("Context", 65), ("Category", 18)]
style_header(ws3, cols3)

# Sort alphabetically by term
index_terms_raw.sort(key=lambda x: x[0].lower())

for i, (term, chapters, context, category) in enumerate(index_terms_raw, 2):
    ws3.cell(row=i, column=1, value=term)
    ws3.cell(row=i, column=2, value=chapters)
    ws3.cell(row=i, column=3, value=context)
    ws3.cell(row=i, column=4, value=category)
    style_row(ws3, i, 4, alt=(i % 2 == 0))

# ── Save ───────────────────────────���─────────────────────────────────────────
output_dir = os.path.join(os.path.dirname(__file__), "output")
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, "book_extraction.xlsx")
wb.save(output_path)
print(f"Created {output_path}")
print(f"  Tab 1 — Citations Needed: {len(citations)} rows")
print(f"  Tab 2 — Cross References: {len(cross_references)} rows")
print(f"  Tab 3 — Index Terms: {len(index_terms_raw)} rows")
